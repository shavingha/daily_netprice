# -*- coding:utf-8 -*-
#This function depends on openpyxl http://openpyxl.readthedocs.io/en/default/
#It is standard python module, you can use "python setup.py install" 

import logging, time, string, urllib2, sys
import openpyxl



def is_stock(cell):
	if len(cell)!=8:
		return False
	prefix = cell[0:2]
	return prefix=="sz" or prefix=="sh"

def init_logging():
	fmt = '%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s'
	datefmt = '%Y-%m-%d %H:%M:%S'
	#display to screen.
	level=logging.DEBUG
	logging.basicConfig(level=level, format=fmt, datefmt=datefmt)
	#logging.basicConfig(level=level, format=fmt, datefmt=datefmt, filename='stock.log', filemode='a')


def get_md(stock_codes):
	if len(stock_codes) == 0:
		return []
	#http://hq.sinajs.cn/list=sh601006,sh600000
	stock_codes = string.join(stock_codes, ",")
	sina_quoteurl = "http://hq.sinajs.cn/list="+stock_codes
	#print(sina_quoteurl)
	uobj = None

	while(uobj is None):
		try:
			uobj = urllib2.urlopen(sina_quoteurl)
		except urllib2.HTTPError, e:
			logging.error(e)
		time.sleep(0.1)
	htmldata = uobj.readlines()
	uobj.close()

	realtime_quotes = {}

	for quote in htmldata:
		#print quote
		tmp_quotes = quote.split("=")
		stock_code = tmp_quotes[0].split("_")[2]
		quote = tmp_quotes[1].strip('";\n')
		if not quote.strip():
			logging.warning(stock_code+" not found.")
			continue
		quote = quote.split(",")
		realtime_quotes[stock_code] = quote

	return realtime_quotes

def get_date():
	import datetime
	now = datetime.datetime.now()
	return now.strftime("%Y-%m-%d")

summary = {'CASH':["CASH", 1, 0]}
def process_acc(rows):
	stock_codes = []
	for row in rows[1:]:
		stock = row[0].value
		#print(stock)
		if stock == "CASH":
			break
		assert(is_stock(stock))
		stock_codes.append(stock)
		
	stock_mds = get_md(stock_codes)
	#print(stock_mds)
	sum_value = 0
	for row in rows[1:]:
		stock = row[0].value

		if stock == "CASH":
			sum_value += row[3].value
			summary["CASH"][2] += row[3].value
			break
		row[1].value = stock_mds[stock][0].decode("GBK")
		row[2].value = string.atof(stock_mds[stock][3])
		sum_value += row[2].value * row[3].value
		stock_name = row[1].value 
		price = row[2].value
		volume = row[3].value
		if summary.has_key(stock):
			cash_array = summary[stock]
			cash_array[2] = cash_array[2] + volume
		else:
			summary[stock] = [stock_name, price, volume]

	row_len = len(rows)
	date = get_date()
	#print(rows[:-1])
	rows[1][5].value = date
	rows[1][6].value = sum_value
def process_sheet(sheet_name):
	if sheet_name[0:3] != "ACC":
		logging.debug("skipping %s", sheet_name)
		return 
	logging.debug("processing sheet:%s", sheet_name)
	ws = wb[sheet_name]
	rows = tuple(ws.rows)
	start_index = -1
	end_index = -1
	i = 0
	for row in rows:
		if row[1].value == "NAME":
			start_index = i
		if row[0].value == "CASH":
			end_index = i
		if (start_index!=-1 and end_index!=-1):
			#print(i, row[1].value)
			#print(start_index, end_index)
			#print(rows[start_index:end_index+1])
			process_acc(rows[start_index:end_index+1])
			start_index = -1
			end_index = -1
			
		i += 1
	#process_acc(rows)


def process_summary(summary):
	ws = wb["SUM"]
	rows = tuple(ws.rows)
	assert(rows[0][0].value == "STOCK")
	i = 3
	cash = summary["CASH"][2]
	ws.cell(row=2,column=4).value = cash
	del summary["CASH"]
	sum = 0
	for k,v in summary.iteritems():
		#	print k,v
		ws.cell(row=i, column=1).value = k
		ws.cell(row=i, column=2).value = v[0]
		ws.cell(row=i, column=3).value = v[1]
		ws.cell(row=i, column=4).value = v[2]
		ws.cell(row=i, column=5).value = v[1]*v[2]
		sum += v[1]*v[2]
		i+=1
	sum+=cash
	
	#clear missed history
	for row in rows[i-1:]:
		for c in row:
			c.value = None
	#ws.garbage_collect()

	#write total summary and date
	date = get_date()
	for row in rows[1:]:
		if row[6].value is None or row[6].value == date:
			row[6].value = date
			row[7].value = sum
			row[8].value = cash
			break
		
	

if __name__ ==  "__main__":
	if len(sys.argv) >= 2:  
		file_name = sys.argv[1]
		init_logging()
		wb = openpyxl.load_workbook(file_name)
		ws_names = wb.get_sheet_names()
		for sheet_name in ws_names:
			process_sheet(sheet_name)
			break

		#print(summary)
		process_summary(summary)
		wb.save(file_name)
	else :
		print "require xlsx file."
